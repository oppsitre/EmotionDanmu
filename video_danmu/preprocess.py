import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import json
import jieba
import os
import matplotlib.pyplot as plt
import math
from sklearn.cluster import KMeans
import numpy as np
from Extract_Audio import get_duration
import config
# from AC_algo_py3_jieba_server import word_segment


def is_chinese(str):
    for uchar in str:
        if not ('\u4e00' <= uchar <= '\u9fff'):
            return False
    return True

# 判断一个unicode是否是数字


def is_number(uchar):
    if '\u0030' <= uchar <= '\u0039':
        return True
    else:
        return False

# 判断一个unicode是否是英文字母


def is_alphabet(uchar):
    if ('\u0041' <= uchar <= '\u005a') or ('\u0061' <= uchar <= '\u007a'):
        return True
    else:
        return False

# 判断是否非汉字，数字和英文字符


def is_other(uchar):
    if not (is_chinese(uchar) or is_number(uchar) or is_alphabet(uchar)):
        return True
    else:
        return False


class Word:
    def __init__(self, content, style):
        self.content = content
        self.style = style


def expressions_read():
    exp_list = []
    with open('../dictionary/expression/expression_labeled_1.txt') as f:
        for line in f:
            line = line.strip().split('\t')
            exp_list.append(line[0])
    return exp_list


def stopwords_read():
    stopwords = []
    with open('../dictionary/word/stopwords.txt') as f:
        for line in f:
            line = line.strip()
            stopwords.append(line)
    return stopwords


def sentence_process(sentence, stopwords):
    sentence_new = []
    for s in sentence:
        if type(s) is tuple:
            sentence_new.append(s[0])
            continue
        seg_list = jieba.cut(s, cut_all=False)
        for seg in seg_list:
            if seg in stopwords:
                continue
            if not is_chinese(seg):
                continue
            sentence_new.append(seg)
    return sentence_new


from openpyxl import load_workbook
emtion2id = {'PA': [0, 0], 'PE': [0, 1], 'PD': [1, 2], 'PH': [1, 3], 'PG':  [1, 4], 'PB': [1, 5], 'PK': [1, 6],
             'NA': [2, 7], 'NB': [3, 8], 'NJ': [3, 9], 'NH': [3, 10], 'PF': [3, 11], 'NI': [4, 12], 'NC': [4, 13],
             'NG': [4, 14], 'NE': [5, 15], 'ND': [5, 16], 'NN': [5, 17], 'NK': [5, 18], 'NL': [5, 19], 'PC': [6, 20]
             }


def emotion_dict_read(files=['../dictionary/word/emotion_dictionary/word_label_1-10000.csv', '../dictionary/word/emotion_dictionary/expression_all.txt']):
    emotion_dict = {}
    with open(files[0], 'r', encoding='utf-8') as f:
        for l in f:
            l = l.strip().split('\t')
            # print(l, flush=True)
            for key in emtion2id.keys():
                if l[3].find(key) != -1:
                    emotion_dict[l[0]] = emtion2id[key]
                    break
    # line = []
    with open(files[1], 'r', encoding='utf-8') as f:
        for l in f:
            l = l.strip().split('\t')
            for key in emtion2id.keys():
                if l[4].find(key) != -1:
                    emotion_dict[l[0]] = emtion2id[key]
                    # print(l[0], key, flush=True)
                    break

    # print('Emotion_Dict', emotion_dict, flush=True)
    return emotion_dict

    # for key in emotion_dict.keys():

    # wb = load_workbook(filename='../dictionary/word/emotion_dic.xlsx') #载入工作簿
    # name2col = { '词语': 'A',
    # 		'词性种类': 'B',
    # 		'词义数': 'C',
    # 		'词义序号': 'D',
    # 		'情感分类': 'E',
    # 		'强度': 'F',
    # 		'极性': 'G',
    # 		'辅助情感分类': 'H',
    # 		'强度': 'I',
    # 		'辅助极性': 'J'
    # 		}
    #
    #
    # ws = wb['Sheet1']  #选中工作表
    # rows = ws.rows
    # columns = ws.columns
    #
    # # 行迭代
    # content = {}
    # for i, row in enumerate(rows):
    # 	if i == 0:
    # 		continue
    # 	line = [col.value for col in row]
    # 	content[line[0]] = [emtion2id[line[4].strip()][1]]
    # return content
    # print(content)
    # print(ws.cell(0,1).value)


def expression_segment(str, expression):
    '''
    The function aims to split word and expressions
    str: the string you want to split
    expression: the list contain all expressions in the dictionary
    Return: a string list and a expression list.
    For example,
    Input: str = 我很开心(^_^)出去玩^_^, expression = [^_^,(^_^),...]
    Output: str_list = ['我很开心', '出去玩'], exp_list = [(^_^), ^_^]
    '''
    pos = 0
    str_tmp = ''
    sentence = []
    while pos < len(str):
        max_len = 0
        exp_tmp = None
        for exp in expression:
            if int(len(str)) - pos < int(len(exp)):
                continue
            if len(exp) <= max_len:
                continue
            if str[pos:(pos + len(exp))] == exp:
                max_len = len(exp)
                exp_tmp = exp
        if max_len > 0:
            if len(str_tmp) > 0:
                sentence.append(str_tmp)
            pos += max_len
            str_tmp = ''
            sentence.append((exp_tmp, 1))

        else:
            str_tmp = str_tmp + str[pos]
            pos += 1

    if len(str_tmp) > 0:
        sentence.append(str_tmp)
    # print('expression_segment')
    return sentence


class Comment:
    def __init__(self, aid, dataset_dir):
        '''
        aid: the av id of each video
        dataset_dir:  the directory of dataset.
        self.content = [[Comment], [Comment], ...]
'''
        self.aid = aid
        self.dataset_dir = dataset_dir
        if not self.dataset_dir.endswith('/'):
            self.dataset_dir += '/'
        self.content = self.read(aid)

    def dfs(self, root):
        cmts = []
        if root is None:
            return None
        for son in root:
            cmts.append(son['content']['message'])
            tmp = self.dfs(son['replies'])
            if tmp != None:
                cmts.extend(tmp)
        return cmts

    def read(self, aid):
        with open(self.dataset_dir + aid + '/' + aid + '.cmt', 'r', encoding='utf-8') as f:
            cmt = json.load(f)
        cmts = self.dfs(cmt['data']['replies'])
        # print(cmts)
        return cmts


class Danmuku:
    def __init__(self, aid, dataset_dir='../dataset/videos/'):
        '''
        aid: the av id of each video
        dataset_dir:  the directory of dataset.
        '''
        self.aid = aid
        self.dataset_dir = dataset_dir
        self.status = None
        if not self.dataset_dir.endswith('/'):
            self.dataset_dir += '/'

        if os.path.exists(dataset_dir + aid + '/' + aid + '.flv'):
            # tmp = get_duration(os.path.abspath(dataset_dir + aid + '/' + aid + '.flv'))
            # if tmp == []:
            #     with open('failed.txt', 'a') as f:
            #         f.write(str(self.aid) + '\n')
            #     self.status = []
            #     return
            self.video_length = get_duration(os.path.abspath(
                dataset_dir + aid + '/' + aid + '.flv'))
        else:
            # print(dataset_dir + aid + '.mp4')
            # tmp = get_duration(os.path.abspath(dataset_dir + aid + '/' + aid + '.mp4'))
            # if tmp == []:
            #     with open('failed.txt', 'a') as f:
            #         f.write(str(self.aid) + '\n')
            #     self.status = []
            #     return
            self.video_length = get_duration(os.path.abspath(
                dataset_dir + aid + '/' + aid + '.mp4'))
        self.content = self.sort(self.read(aid))

        # print(self.content)
        '''
        For example, self.content = [['danmu', offset in the video, the absolute time published],...]
        '''

    def read(self, aid):
        # with open(self.dataset_dir + aid + '/' + aid + '.vid', 'r', encoding = 'utf-8') as f:
        #     vid = json.load(f)
        if self.video_length <= 0:
            print('Video length is 0', aid)
        dan = []
        with open(self.dataset_dir + aid + '/' + aid + '.dan', 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip().split('\t')
                if len(line) != 3 or len(line[-1]) != 19 or line[-1][10] != ' ' or line[-1][4] != '-':
                    continue
                if float(line[1]) >= float(self.video_length):
                    continue
                dan.append((line[0], float(line[1]), line[2]))
        return dan

    def sort(self, content):
        con = sorted(content, key=lambda x: x[1])
        return con

    def cluster(self, content=None, part_num=10):
        if content == None:
            content = self.content
        offset = np.array([c[1] for c in content])
        offset = np.reshape(offset, (len(offset), 1))
        # print('Offset', offset)
        kmeans = KMeans(n_clusters=part_num).fit(offset)
        order = {}
        order_reverse = {}
        idx = 0
        # change the order of labels into increasing order
        for i in kmeans.labels_:
            if not(i in order):
                order[i] = idx
                idx += 1
        labels = [order[item] for idx, item in enumerate(kmeans.labels_)]
        #[Start Time, End Time, Center]
        labels_range = np.ones((part_num, 3)) * -1
        for i in range(len(offset)):
            if labels_range[labels[i], 0] == -1 or offset[i] < labels_range[labels[i], 0]:
                labels_range[labels[i], 0] = offset[i]
            if labels_range[labels[i], 1] == -1 or offset[i] > labels_range[labels[i], 1]:
                labels_range[labels[i], 1] = offset[i]
        centers = sorted(kmeans.cluster_centers_)
        for i in range(part_num):
            labels_range[i, 2] = centers[i]
        # print(centers), labels_range
        # uni, parts = np.unique(labels, return_counts = True)
        # plt.plot(range(part_num), parts)
        # col = ['b', 'g', 'r', 'c', 'm', 'y', 'k']
        # for i in range(len(offset)):
        #     print('I:', i, len(offset), len(labels))
        #     plt.scatter(offset[i], 0, c=col[labels[i]])
        # plt.show()
        return labels, labels_range


def get_all_text(file_video_label='../dataset/texts/video_labeled0.txt', file_video_label_out='../dataset/texts/video_class_2.txt', type='cluster', cluster_num=10):
    # l_tmp = []
    # num = 0
    # with open(file_video_label, 'r', encoding='utf-8') as f:
    #     for line in f:
    #         print(line)
    #         l = line.strip().split('\t')
    #         dan = Danmuku(l[0], '../dataset/videos/')
    #         if len(dan.content) >= cluster_num:
    #             l_tmp.append(line)

    # with open(file_video_label_out, 'w', encoding='utf-8') as f:
    #     for line in l_tmp:
    #         num += 1
    #         f.write(line)

    # print('Total Number of Files', num)
    lines = 0
    num = 0
    # doc_id = open('../dataset/texts/doc2id_1.txt', 'w', encoding = 'utf-8')
    with open('../dataset/texts/texts_1.txt', 'w', encoding='utf-8') as con:
        videos = open(file_video_label, 'r', encoding='utf-8')
        try:
            for line in videos:
                print(line.strip(), flush=True)
                l = line.strip().split('\t')
                list_dirs = os.walk('../dataset/videos/' + l[0])
                for root, dirs, files in list_dirs:
                    for f in files:
                        # print(f, flush=True)
                        if f.endswith('.dan'):
                            dan = Danmuku(l[0], '../dataset/videos/')
                            print('Lines', len(dan.content))
                            lines += len(dan.content)
        except:
            print('Error ID', l[0])
            # labels, labels_range = dan.cluster(part_num = cluster_num)
            # for i in range(cluster_num):
            #     doc_id.write(l[0] + '\t' + str(i+1) + '\t' + str(num) + '\t' + str(labels_range[i, 0]) + '\t' + str(labels_range[i, 1]) + '\t' + str(labels_range[i, 2]) + '\n')
            #     tmp = ''
            #     for j, c in enumerate(dan.content):
            #         if labels[j] == i:
            #             tmp += c[0] + ' '
            #     for j, c in enumerate(tmp):
            #         if c == '\n':
            #             tmp = tmp[:j] + ' ' + tmp[j+1:]
            #     con.write(l[0] + '\t' + str(i+1) + '\t' + str(num) + '\t' + str(labels_range[i, 0]) + '\t' + str(labels_range[i, 1]) + '\t' + str(labels_range[i, 2]) + '\n')
            #     con.write(tmp + '\n')
            #     num += 1
            # break
            # if f.endswith('.cmt'):
            #     cmt = Comment(f[:-4], '../dataset/videos/')
            #     tmp = ''
            #     for c in cmt.content:
            #         tmp += c.strip() + ' '
            #     for j, c in enumerate(tmp):
            #         if c is '\n' or c is '\r' or c is '\r\n':
            #             tmp = tmp[:j] + ' ' + tmp[j+1:]
            #     doc_id.write(f[:-4] + '\t' + '0' + '\t' + str(num) + '\n')
            #     con.write(f[:-4] + '\t' + '0' + '\t' + str(num) + '\n')
            #     con.write(tmp + '\n')
            #     num += 1

    # doc_id.close()
    print('Num:', num)
    print('Total Lines', lines)


def labelfile2id():
    i = 0
    with open('../dataset/texts/video_label2.txt', 'w', encoding='utf-8') as wt:
        with open('../dataset/texts/video_label3.txt', 'r', encoding='utf-8') as f:
            for line in f:
                l = line.strip().split('\t')
                print(i + 1, l[-1])
                i += 1
                for key in emtion2id.keys():
                    if l[-3].find(key) != -1:
                        print(l[-3], emtion2id[key][0], emtion2id[key][1])
                        print(l, key, flus)
                        wt.write('\t'.join(
                            l[:-2]) + '\t' + str(emtion2id[key][0]) + '\t' + str(emtion2id[key][1]) + '\n')
#
# def get_duration(file_name):
#     """get the duration of the time
#     """
#     import re
#     video_info = subprocess.Popen(["ffprobe", file_name],stdout = subprocess.PIPE, stderr = subprocess.STDOUT)
#     video_info  = video_info.stdout.readlines()
#     duration_info = [str(x) for x in video_info if "Duration" in str(x)]
#     duration_str = re.findall(r"Duration: (.+?),", duration_info[0])
#     #print(duration_str)
#     h, m ,s = duration_str[0].split(":")
#     duration = int(h)*3600 + int(m)*60 + float(s)
#
#     return duration


def update_video_duration(root_dir):
    list_dirs = os.walk(root_dir)
    for root, dirs, files in sorted(list_dirs):
        if root[11:] in ['2289679', '4937003']:
            continue
        if root[11:] <= '2274178':
            continue
        vnames = []
        for f in files:
            if f.endswith('.mp4') or f.endswith('.flv'):
                vnames.append(f)
        if len(vnames) == 0:
            continue
        print(root[18:])
        try:
            # name_first = get_first_filename(vnames)
            name_first = vnames[0]
            duration = get_duration(os.path.join(root, name_first))
            print(root[18:], duration)
            aid = root[18:]
            with open(os.path.join(root, aid + '.vid'), 'r', encoding='utf-8') as f:
                vid = json.load(f)
            vid['length'] = int(duration)
            print('VID length', vid['length'])
            with open(os.path.join(root, aid + '.vid'), 'w', encoding='utf-8') as f:
                json.dump(vid, f, ensure_ascii=False, indent=2)
        except (Exception) as e:
            with open('fail_video.txt', 'a') as f:
                print('failed ' + root[18:])
                f.write(root[18:] + '\n')
        # break


def get_first_filename(filenames):
    """get the filename for the multiple videos of one video id
    """
    if len(filenames) == 1:
        return filenames[0]
    nums = []
    len_prefix = 0
    tag = True
    while tag is True:
        for f in filenames:
            if f[len_prefix] != filenames[0][len_prefix]:
                tag = False
        len_prefix += 1
    len_prefix -= 1
    for f in filenames:
        if (f[len_prefix] == '1') and (not is_number(f[len_prefix + 1])):
            return f


e2id = {'快乐': 'PA', '安心': 'PE', '尊敬': 'PD', '赞扬': 'PH', '相信': 'PG', '喜爱': 'PB', '祝愿': 'PK', '愤怒': 'NA',
        '悲伤': 'NB', '失望': 'NJ', '疚': 'NH', '思': 'PF', '慌': 'NI', '恐惧': 'NC', '羞': 'NG', '烦闷': 'NE', '憎恶': 'ND',
        '贬责': 'NN', '妒忌': 'NK', '怀疑': 'NL', '惊奇': 'PC'}


def check_label(filename):
    with open(filename, 'r', encoding='utf-8') as f:
        for l in f:
            l = l.strip().split('\t')
            tag = 0
            for key in e2id.keys():
                if l[5].find(key) != -1:
                    tag += 1
                if l[5].find(e2id[key]) != -1:
                    tag += 1
                if tag % 2 == 1:
                    print(l)
                    break


def main():
    line = []
    with open('../dataset/texts/video_label2.txt', 'r', encoding='utf-8') as f:
        for l in f:
            l = l.strip().split('\t')
            line.append(l)
            print(len(l), l, flush=True)
    with open('../dataset/texts/video_label3.txt', 'w', encoding='utf-8') as f:
        for l in line:
            f.write('\t'.join([l[0], l[1], l[2], l[3], l[5], l[6]]) + '\n')


if __name__ == '__main__':
    # emotion_dict_read()
    # main()
    # labelfile2id()
    # with open('../dataset/texts/video_labeled1.txt', 'r', encoding='utf-8') as f:
    #     for l in f:
    #         print(f)
    # check_label('../dataset/texts/video_labeled1.txt')
    # main()
    # cmt = Comment('4704057', '../dataset/')
    # cmt = Comment('1029329', '../dataset/')
    # l = len(cmt.content)
    # print(cmt.content[:int(l/2)])
    # print(cmt.content[int(l/2):])
    # # print(cmt.content)
    # for c in cmt.content:
    #     print(c)
    #     print(type(c))
    # update_video_duration(root_dir = '../dataset/videos/')
    # f_list = []
    # with open('../dataset/texts/video_labeled2.txt', 'r', encoding = 'utf-8') as f:
    #     for line in f:
    #         line = line.strip().split('\t')
    #         # f_list.append(line[0])
    #         dan = Danmuku(line[0], '../dataset/videos/')
    # get_all_text(cluster_num = config.n_part_danmu)
    # word_segment()
    # labelfile2id()
    # dan = Danmuku('8965800002', '../dataset/videos/')
    # l = len(dan.content)
    # print(dan.video_length)
    # ans, labels_range = dan.cluster(part_num = 10)
    # print(len(dan.content))
    # print(ans)
    # print(labels_range)
    # str = '我很开心^_^出去玩^_^'
    # exp_list = express_read()
    # sentence = expression_extract(str, exp_list)
    # print(sentence)
        # v = Word('我是', 1)
        # print(type(v) is Word)
        # emotion_dic_read()
        # process()
    # wrt = open('output.txt', 'w')
    # with open('test_text.txt', 'r') as f:
    #     tmp = ''
    #     for line in f:
    #         tmp += line
    #     for j, c in enumerate(tmp):
    #         if c == '\n':
    #             tmp = tmp[:j] + ' ' + tmp[j+1:]
    #     wrt.write(tmp + '\n')
    #     print(tmp)
    # with open('../dataset/texts/video_class.txt', 'r', encoding='utf-8') as f:
    #     for l in f:
    #         l = l.strip().split('\t')
    #         dan = Danmuku(l[0])
    #         if len(dan.content) < 10:
    #             print(l, flush=True)
    emotion_dict = emotion_dict_read()
    # print(len(emotion_dict.keys()))
